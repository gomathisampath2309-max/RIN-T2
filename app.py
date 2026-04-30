import streamlit as st
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
from io import BytesIO

# --- PASSWORD PROTECTION ---
st.title("🧪 RespIND T2 Sample Collection Summary")
password = st.text_input("Enter Password:", type="password")
if password != "RIND123":
    st.warning("Please enter the correct password to access data.")
    st.stop()

# --- Load Google Sheet ---
@st.cache_data
def load_sheet(sheet_id):
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv"
    return pd.read_csv(url, on_bad_lines="skip")

df = load_sheet("1Nj-jx92SdX6TOnUXT2QidIXKbdiISFL20euWgrfZPdo")

# --- Data Cleaning ---
df.columns = df.columns.str.strip().str.lower()

# --- Required Columns Check ---
required_cols = ["submissiondate", "sample_id", "type_swab", "p_gender", "p_dob"]
missing = [col for col in required_cols if col not in df.columns]
if missing:
    st.error(f"Missing columns: {missing}")
    st.stop()

# --- FIXED DATE HANDLING (VERY IMPORTANT) ---
df["submissiondate"] = pd.to_datetime(
    df["submissiondate"],
    errors="coerce",
    dayfirst=True  # handles Indian format
)

# Adjust timezone (UTC → IST)
df["submissiondate"] = df["submissiondate"] + pd.Timedelta(hours=5, minutes=30)

# Extract only date
df["date_only"] = df["submissiondate"].dt.date

today = datetime.now().date()

# Filter today
df_today = df[df["date_only"] == today].copy()

# --- DEBUG (remove later if not needed) ---
st.write("Today:", today)
st.write("Available dates:", df["date_only"].dropna().unique())

# --- Sample Type Mapping ---
sample_type_map = {
    "1": "NP swab",
    "2": "OP swab",
    "3": "Nasal swab",
    "4": "ET aspirate",
    "5": "Bronchial lavage /aspirate"
}
df_today["sample_type"] = df_today["type_swab"].astype(str).map(sample_type_map).fillna(df_today["type_swab"])

# --- Age Calculation ---
df_today["p_dob"] = pd.to_datetime(df_today["p_dob"], errors="coerce")

def calculate_age(dob):
    if pd.isna(dob):
        return ""
    today_dt = datetime.today()
    years = today_dt.year - dob.year
    months = today_dt.month - dob.month
    if today_dt.day < dob.day:
        months -= 1
    if months < 0:
        years -= 1
        months += 12
    return f"{years} yr {months} m"

df_today["Age"] = df_today["p_dob"].apply(calculate_age)

# --- Gender Mapping ---
gender_map = {"1": "Male", "2": "Female"}
df_today["Sex"] = df_today["p_gender"].astype(str).map(gender_map).fillna("Other")

# --- Date Column ---
df_today["Date"] = df_today["submissiondate"].dt.strftime("%d-%m-%Y")

# --- Ensure Location Column ---
df_today["location"] = df_today.get("location", "")

# --- Build Table ---
table = df_today[[
    "sample_id",
    "Date",
    "sample_type",
    "p_participant_id",
    "p_child_name",
    "Age",
    "Sex",
    "p_uhid",
    "location",
    "submissiondate"
]].copy()

table.columns = [
    "Sample ID",
    "Date",
    "Sample type",
    "Participant ID",
    "Name",
    "Age",
    "Sex",
    "UHID",
    "Location",
    "Date & time of collection"
]

# --- Add Lab Columns ---
table["Received by"] = ""
table["Volume"] = ""
table["Remarks"] = ""

# --- Add Serial Number ---
table.insert(0, "S.No", range(1, len(table) + 1))

# --- Sort ---
table = table.sort_values(by="Date & time of collection")

# --- Display ---
st.subheader("📋 Today's Sample Collection Details")

if table.empty:
    st.warning("No sample collections found for today.")
else:
    st.dataframe(table, width='stretch')

    # --- Excel Export ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Today_Samples"

    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    align_center = Alignment(horizontal="center", vertical="center")

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=table.shape[1])
    title = ws.cell(row=1, column=1, value="RespIND Daily Sample Collection Summary")
    title.font = Font(bold=True, size=12)
    title.alignment = align_center

    # Header
    for col_num, col_name in enumerate(table.columns, 1):
        cell = ws.cell(row=2, column=col_num, value=col_name)
        cell.font = Font(bold=True)
        cell.alignment = align_center
        cell.border = border

    # Data
    for row_num, row_data in enumerate(table.values, 3):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    st.download_button(
        "⬇️ Download Excel",
        buffer,
        f"{datetime.today().strftime('%d-%m-%Y')}_RespIND_T2_Sample.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
